#! python3
# cellInverter.py 
# This program inverts spreadsheets. ie, what was at row 5, column 3 will
# now be at row 3, column 5

import openpyxl, sys

# Turned this in to a funtion so I can use it in other things
def cell_inverter(target):
    
    try:
        wb = openpyxl.load_workbook(target)
        sheet = wb.active

        # Make a new list and save the spreadsheet data in it as nested lists
        sheet_data = []

        for row in range(sheet.max_row):
            sheet_row = []
            for col in range(sheet.max_column):
                sheet_row.append(sheet.cell(row=row + 1, column=col + 1).value)
            sheet_data.append(sheet_row)
        
        # New workbook for the rotated cells
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active

        for row in range(len(sheet_data)):      
            for col in range(len(sheet_data[row])):
                # Put columns in the rows and rows in the columns
                new_sheet.cell(row=col + 1, column=row + 1).value = sheet_data[row][col]
        
        # Overwrite the original sheet with rotated cells    
        new_wb.save(target)
    
    except FileNotFoundError:
        print("\nThe specified file can't be found")
        sys.exit()

cell_inverter(sys.argv[1])
    