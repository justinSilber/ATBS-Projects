#! python3
# multiplicationTable.py
# This program takes an integer as a command-line argument from the user and 
# generates a multiplication table of n x n dimension

import sys, openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Confirm user has supplied an integer
try:
    size = int(sys.argv[1])

except ValueError:
    print("\nYou must supply an integer")
    sys.exit()

if len(sys.argv) > 2:
    print('\nThank\'s for the extra info, but we\'ll just be using the first integer')

# Create the workbook

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Table'
font_bold = Font(bold = True)

# Write table dimensions and format cells
for num in range(1, size + 1):
    val = num + 1
    sheet.cell(row = 1, column = val).value = num
    sheet.cell(row = 1, column = val).font = font_bold
    sheet.cell(row = val, column = 1).value = num
    sheet.cell(row = val, column = 1).font = font_bold

# Narrow column sizes for readability
for col in range(1, sheet.max_column + 1):
    sheet.column_dimensions[get_column_letter(col)].width = 5

# Multiply the values
for col in range(2, sheet.max_column + 1):
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row = row, column = col).value = sheet.cell(row = 1, column = col).value * sheet.cell(row = row, column = 1).value



wb.save(f'Multiplication Table - {size} x {size}.xlsx')
    